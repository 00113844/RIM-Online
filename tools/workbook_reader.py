"""Read the RIM workbook's *saved* state without needing Excel installed.

openpyxl returns the values Excel cached the last time it recalculated and
saved. For the workbook's own stored scenario those cached values are exact --
full float precision, no display rounding -- so this module can capture a
trustworthy parity fixture on any platform.

It cannot recalculate. Capturing a *new* scenario (different inputs) requires
Excel itself: see tools/excel_oracle.py.
"""
from __future__ import annotations

import glob
import warnings
from pathlib import Path
from typing import Any

import openpyxl

from tools import cell_map as cm

REPO_ROOT = Path(__file__).resolve().parents[1]


def find_workbook(root: Path | None = None) -> Path:
    """Locate the RIM workbook in the repository root."""
    root = root or REPO_ROOT
    matches = sorted(glob.glob(str(root / cm.WORKBOOK_GLOB)))
    if not matches:
        raise FileNotFoundError(
            f"No workbook matching {cm.WORKBOOK_GLOB!r} in {root}. "
            "The .xlsm is gitignored; obtain it from the project owner."
        )
    return Path(matches[0])


def load(path: Path | None = None) -> openpyxl.Workbook:
    """Open the workbook with cached values (formulas are not evaluated)."""
    path = path or find_workbook()
    with warnings.catch_warnings():
        # The workbook uses sparklines, conditional-formatting extensions and
        # drawings that openpyxl warns about and drops. None affect cell values.
        warnings.simplefilter("ignore", UserWarning)
        return openpyxl.load_workbook(path, data_only=True)


def _num(value: Any) -> float | None:
    """Coerce a cell value to float, or None if it is not numeric."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def read_strategy(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read Strategy_X (2.Strategy!D4:M19) as one dict per year.

    Values are the workbook's own labels -- 'Triflur+Tria', 'Narr+B.' and so
    on -- deliberately untranslated. Empty cells become None, which is how
    Calcs!C7:C27 distinguishes an unselected option.
    """
    ws = wb[cm.SHEET_STRATEGY]
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.year_col(year, cm.FIRST_COL_STRATEGY)
        row_data: dict[str, Any] = {"year": year}
        for field, row in cm.STRATEGY_ROWS.items():
            value = ws.cell(row, col).value
            if isinstance(value, str):
                value = value.strip() or None
            row_data[field] = value
        years.append(row_data)
    return years


def read_tabsum(wb: openpyxl.Workbook) -> list[dict[str, float | None]]:
    """Read TabSum (Bio results!C2:M20) -- the within-season state per year.

    Six ryegrass plant stages and ten seed-bank quantities for each of the ten
    years. This is the full biological truth table the Python engine must
    eventually reproduce stage by stage.
    """
    ws = wb[cm.SHEET_BIO]
    stages = {**cm.TABSUM_PLANT_STAGES, **cm.TABSUM_SEED_STAGES}
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.year_col(year, cm.FIRST_COL_BIO)
        years.append(
            {"year": year, **{name: _num(ws.cell(row, col).value) for name, row in stages.items()}}
        )
    return years


def read_ecosum(wb: openpyxl.Workbook) -> tuple[list[dict[str, float | None]], float | None]:
    """Read EcoSum (Eco results!P5:AB17).

    Returns (per-year rows, average gross margin from Eco results!P5).
    Note this block has its own layout: captions in column P, years 1..10 in
    columns Q..Z, totals in AB.
    """
    ws = wb[cm.SHEET_ECO]
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.ECOSUM_FIRST_YEAR_COL + year - 1
        years.append(
            {"year": year, **{name: _num(ws.cell(row, col).value) for name, row in cm.ECOSUM_ROWS.items()}}
        )
    sheet, row, col = cm.ECOSUM_AVERAGE_GM
    return years, _num(wb[sheet].cell(row, col).value)


def read_profile_scenario_name(wb: openpyxl.Workbook) -> str:
    """The active profile label, e.g. 'Susceptible, 2022-LoRF' (2.Strategy!J1)."""
    value = wb[cm.SHEET_STRATEGY].cell(1, 10).value
    return str(value).strip() if value else "unnamed"


def read_rotation_codes(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read the Calcs rows 184-189 rotation cascade for years 1..10.

    This is the workbook's own answer for the coding block, and therefore the
    test oracle for rim/rotation.py.
    """
    ws = wb[cm.SHEET_CALCS]
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.year_col(year, cm.FIRST_COL_ROTATION)
        years.append(
            {"year": year, **{name: _num(ws.cell(row, col).value) for name, row in cm.ROTATION_ROWS.items()}}
        )
    return years


def read_history(wb: openpyxl.Workbook) -> dict[str, str]:
    """Read the paddock history letters from Calcs!N181/N182."""
    out = {}
    for field, (sheet, row, col) in cm.HISTORY_CELLS.items():
        value = wb[sheet].cell(row, col).value
        out[field] = str(value).strip().lower() if value else "w"
    return out


def _raw(value: Any) -> Any:
    """Preserve the blank/number distinction the Calcs block depends on.

    An activation cell holds the crop code when the option is chosen and an
    empty string when it is not, and every downstream formula tests
    ``<> ""`` rather than a numeric value.
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    if isinstance(value, bool):
        return None
    return float(value)


def read_activation(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read Calcs!C7:C49 (per year) -- which control options are active.

    A cell holds the crop code when its option is selected, blank otherwise.
    This is block 2's output; capturing it lets block 3 be ported and tested
    against Excel's own inputs first.
    """
    ws = wb[cm.SHEET_CALCS]
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.year_col(year, cm.FIRST_COL_CALCS)
        years.append(
            {"year": year, **{str(row): _raw(ws.cell(row, col).value) for row in cm.ACTIVATION_ROWS}}
        )
    return years


def read_survival_factors(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read Calcs!C55:C97 (per year) -- the stage survival factors."""
    ws = wb[cm.SHEET_CALCS]
    years = []
    for year in range(1, cm.N_YEARS + 1):
        col = cm.year_col(year, cm.FIRST_COL_CALCS)
        years.append(
            {"year": year, **{str(row): _num(ws.cell(row, col).value) for row in cm.SURVIVAL_ROWS}}
        )
    return years
